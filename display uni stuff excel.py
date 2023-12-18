from openpyxl import Workbook
import openpyxl
import os

import get_uni_info2

filename = 'unis excel file.xlsx'  # change to whatever your excel file is named

if os.path.isfile(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name('Sheet')

else:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Uni name"
    sheet["B1"] = "Rank"
    sheet["C1"] = "Courses"
    sheet["D1"] = "Entry requirement"

    workbook.save(filename=filename)



#l = [["Warwick", {
#        "rank": 5,
#        "courses": [["course 1", "AAA"], ["course 2", "A*AA"]]
#    }], [...]
#]

num_of_unis_inp = int(input("How many unis: "))
start = int(input("Where do you want to start: "))



row = 1

for i in range(start, start+num_of_unis_inp):

    #try:
    uni_info = get_uni_info2.get_info_for_individual(i)
    print(uni_info)


    for a in range(len(uni_info[1]["courses"])):
        sheet.cell(row=row, column=1).value = uni_info[0]
        sheet.cell(row=row, column=2).value = uni_info[1]["rank"]
        sheet.cell(row=row, column=3).value = uni_info[1]["courses"][a][0]
        sheet.cell(row=row, column=4).value = uni_info[1]["courses"][a][1]
        row += 1
    workbook.save("uni stuff.xlsx")

    #except:
    #    row += 1
